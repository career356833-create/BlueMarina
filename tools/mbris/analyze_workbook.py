#!/usr/bin/env python3
"""XLSX 워크북 구조 분석. 원본을 열기만 하고 수정하지 않는다."""
import csv, json, sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
SRC = ROOT / "data" / "mbris" / "raw" / "catalog" / "original" / "mbris-national-species-catalog.xlsx"
OUT = ROOT / "data" / "mbris" / "analysis"
OUT.mkdir(parents=True, exist_ok=True)


def main() -> None:
    # read_only는 병합셀·숨김 정보를 못 읽으므로 일반 모드로 연다
    wb = openpyxl.load_workbook(SRC, data_only=False)
    print(f"시트 {len(wb.sheetnames)}개: {wb.sheetnames}\n")

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))

        # 데이터 시작 행: 비어있지 않은 셀이 2개 이상인 첫 행을 헤더로 본다
        header_idx, header = None, []
        for i, row in enumerate(rows[:20]):
            filled = [c for c in row if c not in (None, "")]
            if len(filled) >= 2:
                header_idx = i
                header = [str(c).strip() if c is not None else "" for c in row]
                break

        data_rows = rows[header_idx + 1:] if header_idx is not None else []
        non_empty = [r for r in data_rows if any(c not in (None, "") for c in r)]

        # 수식 탐지 (앞 200행만)
        formulas = sum(
            1 for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200))
            for c in r if isinstance(c.value, str) and c.value.startswith("="))

        dup_header = [h for h in set(header) if h and header.count(h) > 1]

        info = {
            "sheetName": name,
            "state": ws.sheet_state,          # visible / hidden / veryHidden
            "maxRow": ws.max_row,
            "maxColumn": ws.max_column,
            "headerRowIndex": (header_idx + 1) if header_idx is not None else None,
            "dataStartRow": (header_idx + 2) if header_idx is not None else None,
            "header": [h for h in header if h],
            "headerRaw": header,
            "duplicateHeaders": dup_header,
            "dataRowCount": len(non_empty),
            "emptyRowCount": len(data_rows) - len(non_empty),
            "mergedCellCount": len(ws.merged_cells.ranges),
            "mergedCellSample": [str(x) for x in list(ws.merged_cells.ranges)[:5]],
            "formulaCount": formulas,
            "hasAutoFilter": ws.auto_filter.ref is not None,
            "autoFilterRef": ws.auto_filter.ref,
            "tableCount": len(getattr(ws, "tables", {}) or {}),
            "freezePanes": ws.freeze_panes,
        }
        sheets.append(info)

        print(f"[{name}] state={ws.sheet_state} rows={ws.max_row} cols={ws.max_column} "
              f"data={len(non_empty)} merged={info['mergedCellCount']} formula={formulas}")
        print(f"   헤더행={info['headerRowIndex']}  컬럼({len(info['header'])}): {info['header']}")
        if dup_header:
            print(f"   ⚠️ 중복 헤더: {dup_header}")
        # 첫 데이터 행 샘플
        for r in non_empty[:2]:
            print(f"   샘플: {[str(c)[:22] if c is not None else None for c in r]}")
        print()

    result = {
        "sourceFile": SRC.name,
        "analyzedAt": datetime.now(timezone.utc).isoformat(),
        "sheetCount": len(sheets),
        "totalDataRows": sum(s["dataRowCount"] for s in sheets),
        "sheets": sheets,
    }
    (OUT / "workbook-structure.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    with (OUT / "sheet-summary.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheetName", "state", "maxRow", "maxColumn", "headerRow",
                    "dataRowCount", "emptyRowCount", "mergedCells", "formulas",
                    "duplicateHeaders", "columns"])
        for s in sheets:
            w.writerow([s["sheetName"], s["state"], s["maxRow"], s["maxColumn"],
                        s["headerRowIndex"], s["dataRowCount"], s["emptyRowCount"],
                        s["mergedCellCount"], s["formulaCount"],
                        " | ".join(s["duplicateHeaders"]), " | ".join(s["header"])])

    print(f"총 데이터 행: {result['totalDataRows']:,}")
    print(f"저장: {OUT}")


if __name__ == "__main__":
    main()
